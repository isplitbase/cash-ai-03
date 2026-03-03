import os
import sys
import traceback
from pathlib import Path

import openpyxl


def _work_dir() -> Path:
    # runner 側で WORK_DIR を渡す。未指定ならカレントを採用。
    return Path(os.environ.get("WORK_DIR", ".")).resolve()


def _truthy_env(name: str, default: bool = False) -> bool:
    v = os.environ.get(name)
    if v is None:
        return default
    return v.strip().lower() in {"1", "true", "yes", "y", "on"}


def copy_formulas_or_values(src_sheet, dest_sheet, range_str: str) -> None:
    """
    指定された範囲の「数式」または「値」をコピーする。

    - src の読み込みを data_only=False にすると cell.value は数式文字列（例: '=SUM(A1:A5)'）になる
    - data_only=True にすると cell.value は（保存されている）計算結果の値になる

    結合セルの場合は、その結合範囲の左上セルに対してのみ書き込みを行う。
    """
    # 結合セルの範囲一覧（コピー先）を取得
    merged_ranges = list(dest_sheet.merged_cells.ranges)

    for row in src_sheet[range_str]:
        for cell in row:
            val = cell.value
            if val is None:
                continue

            coord = cell.coordinate

            # coord が結合セル範囲内なら、左上セル座標に寄せる
            target_coord = coord
            is_merged_child = False

            for r in merged_ranges:
                if coord in r:
                    top_left_coord = openpyxl.utils.get_column_letter(r.min_col) + str(r.min_row)
                    if coord != top_left_coord:
                        is_merged_child = True
                    target_coord = top_left_coord
                    break

            # 結合セルの枝（左上以外）には書き込まない
            if is_merged_child:
                continue

            dest_sheet[target_coord].value = val


def main() -> None:
    print("colab1-5.py: 処理を開始します（数式/値コピー・結合セル完全対応版）...")

    work_dir = _work_dir()

    # 入出力ファイル（runner201.py と同じ WORK_DIR 配下）
    file1_path = work_dir / "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx"
    file2_path = work_dir / "CF資金移動表.xlsx"
    out_path = work_dir / "CF資金移動表_updated.xlsx"

    print(f"WORK_DIR = {work_dir}")
    print(f"INPUT    = {file1_path}")
    print(f"TEMPLATE = {file2_path}")
    print(f"OUTPUT   = {out_path}")

    if not file1_path.exists():
        raise FileNotFoundError(f"入力が見つかりません: {file1_path}")
    if not file2_path.exists():
        raise FileNotFoundError(
            f"テンプレが見つかりません: {file2_path}（app/pipeline/assets に同梱してください）"
        )

    # 互換性のため、デフォルトは「値コピー（従来colab1-5.py相当）」にしておく。
    # 新ロジックで「数式をコピーしたい」場合は COPY_FORMULAS=1 を渡す。
    copy_formulas = _truthy_env("COPY_FORMULAS", default=False)

    # ワークブックの読み込み
    wb_src = openpyxl.load_workbook(str(file1_path), data_only=(not copy_formulas))
    wb_dest = openpyxl.load_workbook(str(file2_path))

    print(f"MODE = {'FORMULAS' if copy_formulas else 'VALUES'}")
    print(f"SRC sheets : {wb_src.sheetnames}")
    print(f"DEST sheets: {wb_dest.sheetnames}")

    # コピー対象の設定
    copy_tasks = [
        ("財務諸表（入力）", "A4:O185"),
        ("資金移動表", "A7:O73"),
        ("CF計算書", "B9:C50"),
        ("CF計算書②", "B9:C50"),
    ]

    for sheet_name, cell_range in copy_tasks:
        if sheet_name in wb_src.sheetnames and sheet_name in wb_dest.sheetnames:
            print(f"コピー中: {sheet_name} ({cell_range})")
            src_ws = wb_src[sheet_name]
            dest_ws = wb_dest[sheet_name]
            copy_formulas_or_values(src_ws, dest_ws, cell_range)
        else:
            print(f"警告: シート '{sheet_name}' が見つからないためスキップしました。")

    # 保存（更新版として別名で保存）
    wb_dest.save(str(out_path))
    print(f"colab1-5.py: 完了しました。保存先: {out_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        # Cloud Run の stderr にフルスタックを出す（原因特定用）
        traceback.print_exc(file=sys.stderr)
        raise
