# -*- coding: utf-8 -*-
"""
colab201-ipan.py
JSON -> Excel 転記（仕様駆動）
  + LibreOfficeで再計算（headless）
  + ★全シートで「数式 → 結果（値貼り）」へ置換
  + 追加仕様：勘定科目が空の場合、集計方法も必ず空にする

前提（WORK_DIR 配下）:
  - エクセル転記仕様.xlsx
  - output_updated.json
  - CF付財務分析表（経営指標あり）_ReadingData.xlsx

出力（WORK_DIR 配下）:
  - CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx
  - transfer_log.txt
"""

import json
import re
import math
import os
import sys
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


# ==========
# パス設定（WORK_DIR 前提に統一）
# ==========
BASE_DIR = Path(os.environ.get("WORK_DIR", "/tmp/work")).resolve()

SPEC_PATH = BASE_DIR / "エクセル転記仕様.xlsx"
SRC_EXCEL_PATH = BASE_DIR / "CF付財務分析表（経営指標あり）_ReadingData.xlsx"
JSON_PATH = BASE_DIR / "output_updated.json"

OUT_EXCEL_PATH = BASE_DIR / "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx"
OUT_LOG_PATH = BASE_DIR / "transfer_log.txt"

TARGET_SHEET_NAME = "財務諸表（入力）"


# ==========
# 行集合パース
# ==========
def parse_row_set(expr: str) -> set[int]:
    if expr is None:
        return set()
    s = str(expr).strip()
    if not s:
        return set()
    s = s.replace(" ", "").replace("　", "")
    parts = [p for p in s.split(",") if p]
    rows: set[int] = set()
    for p in parts:
        if "-" in p:
            a, b = p.split("-", 1)
            if a == "" or b == "":
                raise ValueError(f"Invalid range token: {p}")
            a_i = int(a)
            b_i = int(b)
            if a_i > b_i:
                raise ValueError(f"Range start > end: {p}")
            rows.update(range(a_i, b_i + 1))
        else:
            rows.add(int(p))
    return rows


# ==========
# 値変換（JSON -> Excel）
# ==========
def coerce_value(v):
    if v is None:
        return None

    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v

    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None

        # 数値文字列対応（カンマ除去）
        s2 = s.replace(",", "").replace("，", "")

        # (123) -> -123
        m = re.fullmatch(r"\(([-+]?\d+(\.\d+)?)\)", s2)
        if m:
            s2 = "-" + m.group(1)

        try:
            if re.fullmatch(r"[-+]?\d+", s2):
                return int(s2)
            return float(s2)
        except Exception:
            return s

    return v


# ==========
# 結合セル対応
# ==========
def writable_cell(ws, row: int, col_letter: str):
    col_idx = column_index_from_string(col_letter)
    requested_cell = ws.cell(row=row, column=col_idx)
    coord = f"{col_letter}{row}"

    if not ws.merged_cells.ranges:
        return requested_cell

    for mr in ws.merged_cells.ranges:
        if coord in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col)

    return requested_cell


def is_merged_child(ws, coord: str) -> bool:
    """coord が結合セルの“左上以外”なら True"""
    for mr in ws.merged_cells.ranges:
        if coord in mr:
            top_left = f"{get_column_letter(mr.min_col)}{mr.min_row}"
            return coord != top_left
    return False


# ==========
# Excel起動時フル再計算設定（保険）
# ==========
def set_recalc_on_load(workbook: openpyxl.Workbook):
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True


# ==========
# LibreOfficeで再計算（xlsx -> xlsx 変換）
# ==========
def libreoffice_recalc_xlsx(src_xlsx: Path, out_dir: Path, log_lines: list[str]) -> Path:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError("LibreOffice が見つかりません（soffice/libreoffice コマンドが必要です）")

    out_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        soffice,
        "--headless",
        "--nologo",
        "--nolockcheck",
        "--nodefault",
        "--norestore",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(out_dir),
        str(src_xlsx),
    ]
    log_lines.append(f"[LO] run: {' '.join(cmd)}")

    p = subprocess.run(cmd, capture_output=True, text=True)
    log_lines.append(f"[LO] returncode={p.returncode}")
    if p.stdout.strip():
        log_lines.append(f"[LO] stdout: {p.stdout.strip()}")
    if p.stderr.strip():
        log_lines.append(f"[LO] stderr: {p.stderr.strip()}")

    if p.returncode != 0:
        raise RuntimeError("LibreOffice 変換（再計算）に失敗しました。stderr を確認してください。")

    out_path = out_dir / src_xlsx.name
    if not out_path.exists():
        # basename違いの保険
        candidates = list(out_dir.glob(src_xlsx.stem + "*.xlsx"))
        if len(candidates) == 1:
            out_path = candidates[0]
        else:
            raise RuntimeError(f"LibreOffice 変換結果が見つかりません: {out_path}")

    return out_path


# ==========
# ★数式→値貼り（全シート）
# ==========
def replace_formulas_with_values_from_cache_all_sheets(
    wb_formula: openpyxl.Workbook,
    wb_values: openpyxl.Workbook,
    log_lines: list[str],
) -> None:
    replaced = 0
    skipped_merged_children = 0
    skipped_missing_sheet = 0

    for name in wb_formula.sheetnames:  # ★全シート
        if name not in wb_values.sheetnames:
            skipped_missing_sheet += 1
            continue

        ws_f = wb_formula[name]
        ws_v = wb_values[name]

        for row in ws_f.iter_rows():
            for cell_f in row:
                v = cell_f.value
                is_formula = (cell_f.data_type == "f") or (isinstance(v, str) and v.startswith("="))
                if not is_formula:
                    continue

                coord = cell_f.coordinate
                if is_merged_child(ws_f, coord):
                    skipped_merged_children += 1
                    continue

                # data_only=True 側の同セルの値（計算結果）を貼る
                cell_f.value = ws_v[coord].value
                replaced += 1

    log_lines.append(
        f"[PASTE] replaced_formula_cells={replaced}, "
        f"skipped_merged_children={skipped_merged_children}, "
        f"skipped_missing_sheet={skipped_missing_sheet}"
    )


# ==========
# 仕様（転記ルール）読み取り：ヘッダ自動検出
# ==========
def find_header_row_and_cols(ws):
    max_scan = min(ws.max_row, 100)
    for r in range(1, max_scan + 1):
        colmap = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s in ("対象", "キー名", "転記列", "転記行"):
                colmap[s] = c
        if ("対象" in colmap) and ("転記列" in colmap) and ("転記行" in colmap):
            return r, colmap
    return None, None


def load_rules(spec_path: Path):
    spec_wb = openpyxl.load_workbook(str(spec_path), data_only=True)

    # よくあるシート名を優先
    sheet_order = []
    for n in ("ルール(正)", "Sheet2", "Sheet1"):
        if n in spec_wb.sheetnames:
            sheet_order.append(n)
    for n in spec_wb.sheetnames:
        if n not in sheet_order:
            sheet_order.append(n)

    for sheet_name in sheet_order:
        ws = spec_wb[sheet_name]
        header_row, colmap = find_header_row_and_cols(ws)
        if header_row is None:
            continue

        has_key = "キー名" in colmap

        # キー名が無い仕様の場合のフォールバック
        target_to_key = {
            "区分": "category",
            "勘定科目": "account_name",
            "前々期": "value_t-2",
            "前期": "value_t-1",
            "今期": "value_t",
            "備考": "aggregation_method",
            "集計方法": "aggregation_method",
        }

        rules = []
        for r in range(header_row + 1, ws.max_row + 1):
            target = ws.cell(r, colmap["対象"]).value
            col = ws.cell(r, colmap["転記列"]).value
            row_expr = ws.cell(r, colmap["転記行"]).value
            key_name = ws.cell(r, colmap["キー名"]).value if has_key else None

            if target is None and col is None and row_expr is None and (not has_key or key_name is None):
                continue
            if target is None or col is None or row_expr is None:
                continue

            target_s = str(target).strip()
            col_s = str(col).strip()
            allowed_rows = parse_row_set(str(row_expr))

            if has_key:
                if key_name is None:
                    continue
                key_s = str(key_name).strip()
            else:
                key_s = target_to_key.get(target_s)
                if key_s is None:
                    continue

            rules.append({
                "target": target_s,
                "key_name": key_s,         # 仕様上のキー名（内部キー）
                "column": col_s,           # 例: "B"
                "allowed_rows": allowed_rows,
            })

        if rules:
            return rules, sheet_name, header_row

    raise ValueError("転記ルールが見つかりません。仕様シートに '対象/転記列/転記行' のヘッダ行が必要です。")


# ==========
# メイン
# ==========
def main() -> int:
    log_lines: list[str] = []
    log_lines.append("=== Excel Transfer Log ===")
    log_lines.append(f"Timestamp: {datetime.now().isoformat(timespec='seconds')}")
    log_lines.append(f"WORK_DIR: {BASE_DIR}")
    log_lines.append(f"Spec: {SPEC_PATH}")
    log_lines.append(f"Source Excel: {SRC_EXCEL_PATH}")
    log_lines.append(f"JSON: {JSON_PATH}")
    log_lines.append(f"Target Sheet: {TARGET_SHEET_NAME}")
    log_lines.append("")

    try:
        # 必須ファイル確認
        for p in [SPEC_PATH, SRC_EXCEL_PATH, JSON_PATH]:
            if not p.exists():
                raise FileNotFoundError(f"Required file not found: {p}")

        # 仕様読み取り
        rules, rules_sheet, header_row = load_rules(SPEC_PATH)
        log_lines.append(f"[SPEC] sheet={rules_sheet} header_row={header_row} rules={len(rules)}")

        # JSON読み取り
        with open(JSON_PATH, "r", encoding="utf-8") as f:
            records = json.load(f)
        if not isinstance(records, list):
            raise ValueError("output_updated.json must be a JSON array (list) of objects.")
        log_lines.append(f"[JSON] records={len(records)}")

        # Excel読み取り（数式保持）
        wb = openpyxl.load_workbook(str(SRC_EXCEL_PATH), data_only=False)
        if TARGET_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f'Sheet "{TARGET_SHEET_NAME}" not found in {SRC_EXCEL_PATH}')
        ws = wb[TARGET_SHEET_NAME]

        # 再計算フラグ（保険）
        set_recalc_on_load(wb)

        # 仕様キー名 → JSONキー名
        SPEC_TO_JSON_KEY = {
            "category": "区分",
            "account_name": "勘定科目",
            "value_t-2": "前々期",
            "value_t-1": "前期",
            "value_t": "今期",
            "aggregation_method": "集計方法",
            "remark": "集計方法",
        }

        stats = {
            "records_total": len(records),
            "records_used": 0,
            "rows_written": 0,
            "cells_written": 0,
            "skipped_sheet_mismatch": 0,
            "skipped_no_cell": 0,
            "skipped_bad_cell": 0,
            "skipped_not_allowed": 0,
            "missing_key": 0,
            "writes_to_merged_anchor": 0,
            "remark_forced_blank": 0,
        }

        for rec in records:
            if not isinstance(rec, dict):
                continue

            rec_sheet = rec.get("シート名")
            if rec_sheet is not None and rec_sheet != TARGET_SHEET_NAME:
                stats["skipped_sheet_mismatch"] += 1
                continue

            cell_row_raw = rec.get("セル")
            if cell_row_raw is None:
                stats["skipped_no_cell"] += 1
                continue

            try:
                excel_row = int(str(cell_row_raw).strip())
            except Exception:
                stats["skipped_bad_cell"] += 1
                continue

            account_val = coerce_value(rec.get("勘定科目"))
            account_is_blank = (account_val is None) or (isinstance(account_val, str) and account_val.strip() == "")

            stats["records_used"] += 1
            wrote_any = False

            for rule in rules:
                key_name = rule["key_name"]
                col_letter = rule["column"]
                allowed_rows = rule["allowed_rows"]

                if excel_row not in allowed_rows:
                    stats["skipped_not_allowed"] += 1
                    continue

                json_key = SPEC_TO_JSON_KEY.get(key_name, key_name)

                # 追加仕様：勘定科目が空なら集計方法も空
                if json_key == "集計方法" and account_is_blank:
                    value = None
                    stats["remark_forced_blank"] += 1
                else:
                    if json_key not in rec:
                        stats["missing_key"] += 1
                        continue
                    value = coerce_value(rec.get(json_key))

                requested_coord = f"{col_letter}{excel_row}"
                cell = writable_cell(ws, excel_row, col_letter)
                if cell.coordinate != requested_coord:
                    stats["writes_to_merged_anchor"] += 1

                cell.value = value
                stats["cells_written"] += 1
                wrote_any = True

            if wrote_any:
                stats["rows_written"] += 1

        # いったん保存（数式は残る）
        OUT_EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(OUT_EXCEL_PATH))
        log_lines.append(f"[SAVE] formula_kept: {OUT_EXCEL_PATH}")

        # LibreOfficeで再計算→値貼り（全シート）
        with tempfile.TemporaryDirectory(prefix="lo_recalc_", dir=str(BASE_DIR)) as td:
            lo_dir = Path(td)
            evaluated = libreoffice_recalc_xlsx(OUT_EXCEL_PATH, lo_dir, log_lines)

            wb_formula = openpyxl.load_workbook(str(OUT_EXCEL_PATH), data_only=False)
            wb_values = openpyxl.load_workbook(str(evaluated), data_only=True)

            replace_formulas_with_values_from_cache_all_sheets(wb_formula, wb_values, log_lines)
            set_recalc_on_load(wb_formula)
            wb_formula.save(str(OUT_EXCEL_PATH))
            log_lines.append(f"[SAVE] values_pasted: {OUT_EXCEL_PATH}")

        # ログ
        log_lines.append("")
        log_lines.append("=== Summary ===")
        for k, v in stats.items():
            log_lines.append(f"{k}: {v}")

        OUT_LOG_PATH.write_text("\n".join(log_lines), encoding="utf-8")

        print("DONE")
        print("Final Excel:", str(OUT_EXCEL_PATH))
        print("Log:", str(OUT_LOG_PATH))
        return 0

    except Exception as e:
        # 失敗時もログを残す
        log_lines.append("")
        log_lines.append("=== ERROR ===")
        log_lines.append(str(e))
        try:
            OUT_LOG_PATH.write_text("\n".join(log_lines), encoding="utf-8")
        except Exception:
            pass
        print(str(e), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
