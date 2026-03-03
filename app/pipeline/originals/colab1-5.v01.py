import os
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell


def _work_dir() -> Path:
    # runner 側で WORK_DIR を渡す。未指定ならカレントを採用。
    return Path(os.environ.get("WORK_DIR", ".")).resolve()


def copy_values(src_sheet, dest_sheet, range_str: str) -> None:
    """
    指定された範囲の値をコピーする。
    コピー先が結合セルの場合は、その範囲の「左上」のセルのみ書き込みを許可する。
    """
    for row in src_sheet[range_str]:
        for cell in row:
            dest_cell = dest_sheet[cell.coordinate]
            if not isinstance(dest_cell, MergedCell):
                dest_cell.value = cell.value


def main() -> None:
    print("colab1-5.py: 処理を開始します（結合セル対策済み）...")

    work_dir = _work_dir()

    # 入出力ファイル（runner201.py と同じ WORK_DIR 配下）
    file1_path = work_dir / "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx"
    file2_path = work_dir / "CF資金移動表.xlsx"
    out_path = work_dir / "CF資金移動表_updated.xlsx"

    if not file1_path.exists():
        raise FileNotFoundError(f"入力が見つかりません: {file1_path}")
    if not file2_path.exists():
        raise FileNotFoundError(
            f"テンプレが見つかりません: {file2_path}（app/pipeline/assets に同梱してください）"
        )

    # ワークブックの読み込み
    wb_src = openpyxl.load_workbook(str(file1_path), data_only=True)
    wb_dest = openpyxl.load_workbook(str(file2_path))

    # コピー対象の設定
    copy_tasks = [
        ("財務諸表（入力）", "A4:O185"),
        ("資金移動表", "A7:O73"),
        ("CF計算書", "B9:C50"),
        ("CF計算書②", "B9:C50"),
    ]

    for sheet_name, cell_range in copy_tasks:
        if sheet_name in wb_src.sheetnames and sheet_name in wb_dest.sheetnames:
            print(f"コピー中: {sheet_name} ({cell_range})")
            src_ws = wb_src[sheet_name]
            dest_ws = wb_dest[sheet_name]
            copy_values(src_ws, dest_ws, cell_range)
        else:
            print(f"警告: シート '{sheet_name}' が見つからないためスキップしました。")

    # 保存（更新版として別名で保存）
    wb_dest.save(str(out_path))
    print(f"colab1-5.py: 完了しました。保存先: {out_path}")


if __name__ == "__main__":
    main()
