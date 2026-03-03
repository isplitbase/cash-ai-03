# -*- coding: utf-8 -*-
"""
水野さんの１－３
JSON -> Excel 転記（「財務諸表（入力）」シート限定版）

前提（WORK_DIR 配下）:
  - エクセル転記仕様.xlsx
  - output_updated.json
  - CF付財務分析表（経営指標あり）_ReadingData.xlsx

出力（WORK_DIR 配下）:
  - CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx
  - transfer_log.txt
"""

import json
import re
import math
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string

# ==========
# パス設定（cash-ai-03 runner の前提に合わせる）
# ==========
BASE_DIR = os.environ.get("WORK_DIR", "/tmp/work")

SPEC_PATH = os.path.join(BASE_DIR, "エクセル転記仕様.xlsx")
SRC_EXCEL_PATH = os.path.join(BASE_DIR, "CF付財務分析表（経営指標あり）_ReadingData.xlsx")
JSON_PATH = os.path.join(BASE_DIR, "output_updated.json")

OUT_EXCEL_PATH = os.path.join(BASE_DIR, "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx")
OUT_LOG_PATH = os.path.join(BASE_DIR, "transfer_log.txt")

TARGET_SHEET_NAME = "財務諸表（入力）"


# ==========
# 補助関数
# ==========

def parse_row_set(expr: str) -> set[int]:
    """「転記行」のパース（例: '6-10,12'）"""
    if expr is None:
        return set()
    s = str(expr).strip().replace(" ", "").replace("　", "")
    if not s:
        return set()
    parts = [p for p in s.split(",") if p]
    rows: set[int] = set()
    for p in parts:
        if "-" in p:
            a, b = p.split("-", 1)
            rows.update(range(int(a), int(b) + 1))
        else:
            rows.add(int(p))
    return rows


def coerce_value(v):
    """JSONの値をExcel用に変換（数値変換・カンマ除去）"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and (math.isnan(v) or math.isinf(v))) else v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s2 = s.replace(",", "").replace("，", "")
        # (123) -> -123 対応
        m = re.fullmatch(r"\(([-+]?\d+(\.\d+)?)\)", s2)
        if m:
            s2 = "-" + m.group(1)
        try:
            if re.fullmatch(r"[-+]?\d+", s2):
                return int(s2)
            return float(s2)
        except:
            return s
    return v


def writable_cell(ws, row: int, col_letter: str):
    """結合セルの場合、左上（アンカー）セルを返す"""
    col_idx = column_index_from_string(col_letter)
    coord = f"{col_letter}{row}"
    for mr in ws.merged_cells.ranges:
        if coord in mr:
            return ws.cell(row=mr.bounds[1], column=mr.bounds[0])
    return ws.cell(row=row, column=col_idx)


# ==========
# メイン処理
# ==========

def main() -> int:
    # 0) ファイル存在チェック（runner 側が returncode を見るので、失敗時は非0で終了）
    for p in [SPEC_PATH, SRC_EXCEL_PATH, JSON_PATH]:
        if not os.path.exists(p):
            print(f"エラー: ファイルが見つかりません -> {p}", file=sys.stderr)
            return 1

    # 1) 転記仕様の読み取り
    spec_wb = openpyxl.load_workbook(SPEC_PATH, data_only=True)
    sheet_name = "ルール(正)" if "ルール(正)" in spec_wb.sheetnames else spec_wb.sheetnames[0]
    rules_ws = spec_wb[sheet_name]

    rules = []
    # 3行目から読み取り（ヘッダが2行目想定）
    for r in range(3, rules_ws.max_row + 1):
        target = rules_ws.cell(r, 1).value
        key_name = rules_ws.cell(r, 2).value
        col = rules_ws.cell(r, 3).value
        row_expr = rules_ws.cell(r, 4).value
        if not all([target, key_name, col, row_expr]):
            continue

        rules.append({
            "key_name": str(key_name),
            "column": str(col).strip(),
            "allowed_rows": parse_row_set(str(row_expr)),
        })

    # 2) JSON読み取り
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        records = json.load(f)

    # 3) Excel読み取り
    wb = openpyxl.load_workbook(SRC_EXCEL_PATH)
    if TARGET_SHEET_NAME not in wb.sheetnames:
        print(f"エラー: シート「{TARGET_SHEET_NAME}」がありません。", file=sys.stderr)
        return 1
    ws = wb[TARGET_SHEET_NAME]

    # Excelを開いた時に再計算させる設定（元コード踏襲）
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # 4) 転記実行
    log_lines = [f"--- Transfer Log {datetime.now()} ---"]
    stats = {"cells_written": 0, "rows_written": 0}

    SPEC_TO_JSON_KEY = {
        "category": "区分",
        "account_name": "勘定科目",
        "value_t-2": "前々期",
        "value_t-1": "前期",
        "value_t": "今期",
        "aggregation_method": "集計方法",
    }

    for rec in records:
        cell_row_raw = rec.get("セル")
        if cell_row_raw is None:
            continue
        try:
            excel_row = int(str(cell_row_raw).strip())
        except:
            continue

        # 追加仕様：勘定科目が空なら集計方法も空にする
        account_val = coerce_value(rec.get("勘定科目"))
        account_is_blank = (account_val is None or str(account_val).strip() == "")

        wrote_any = False
        for rule in rules:
            key_name = rule["key_name"]
            if excel_row not in rule["allowed_rows"]:
                continue

            json_key = SPEC_TO_JSON_KEY.get(key_name, key_name)

            # 勘定科目が空の場合の「集計方法」空処理
            if (json_key == "集計方法" or json_key == "備考") and account_is_blank:
                value = None
            else:
                value = coerce_value(rec.get(json_key))

            cell = writable_cell(ws, excel_row, rule["column"])
            cell.value = value
            stats["cells_written"] += 1
            wrote_any = True

        if wrote_any:
            stats["rows_written"] += 1

    # 5) 保存
    wb.save(OUT_EXCEL_PATH)

    # 6) ログ出力
    log_lines.append(f"Summary: {stats}")
    with open(OUT_LOG_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))

    print("完了しました。")
    print(f"出力ファイル: {OUT_EXCEL_PATH}")
    print(f"書き込みセル数: {stats['cells_written']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
