import os
from pathlib import Path

import openpyxl


def _work_dir() -> Path:
    # runner 側で WORK_DIR を渡す。未指定ならカレントを採用。
    return Path(os.environ.get("WORK_DIR", ".")).resolve()


def copy_formulas_or_values(src_sheet, dest_sheet, range_str: str) -> None:
    """
    指定された範囲の「数式」または「値」をコピーする。
    結合セルの場合は、その範囲の左上セルに対してのみ書き込みを行う。
    """
    merged_ranges = list(dest_sheet.merged_cells.ranges)

    for row in src_sheet[range_str]:
        for cell in row:
            val = cell.value
            if val is None:
                continue

            coord = cell.coordinate

            # コピー先のセルが結合セルの一部かどうかを判定し、
            # 結合範囲の左上セルだけに書き込む
            target_coord = coord
            is_merged_child = False

            for r in merged_ranges:
                if coord in r:
                    top_left_coord = openpyxl.utils.get_column_letter(r.min_col) + str(r.min_row)
                    if coord != top_left_coord:
                        is_merged_child = True
                    target_coord = top_left_coord
                    break

            if not is_merged_child:
                dest_sheet[target_coord].value = val


def main() -> None:
    print("colab1-5.py: 処理を開始します（数式コピー・結合セル完全対応版）...")

    work_dir = _work_dir()

    # ★ colab1-5.py と同じ入出力に合わせる
    file1_path = work_dir / "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx"
    file2_path = work_dir / "CF資金移動表.xlsx"
    out_path = work_dir / "CF資金移動表_updated.xlsx"

    if not file1_path.exists():
        raise FileNotFoundError(f"入力が見つかりません: {file1_path}")
    if not file2_path.exists():
        raise FileNotFoundError(
            f"テンプレが見つかりません: {file2_path}（app/pipeline/assets に同梱してください）"
        )

    # ★ 新ロジック（colab1-5.new.py）に合わせて data_only=False で数式も取得
    wb_src = openpyxl.load_workbook(str(file1_path), data_only=False)
    wb_dest = openpyxl.load_workbook(str(file2_path))

    copy_tasks = [
        ("財務諸表（入力）", "A4:O185"),
        ("資金移動表", "A7:O73"),
        ("CF計算書", "B9:C50"),
        ("CF計算書②", "B9:C50"),
    ]

    for sheet_name, cell_range in copy_tasks:
        if sheet_name in wb_src.sheetnames and sheet_name in wb_dest.sheetnames:
            print(f"数式コピー中: {sheet_name} ({cell_range})")
            src_ws = wb_src[sheet_name]
            dest_ws = wb_dest[sheet_name]
            copy_formulas_or_values(src_ws, dest_ws, cell_range)
        else:
            print(f"警告: シート '{sheet_name}' が見つからないためス_
