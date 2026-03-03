# -*- coding: utf-8 -*-

import json
import math
from typing import List, Dict, Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

from openai import OpenAI

# =========================
# 設定
# =========================

import os

BASE_DIR = os.environ.get("WORK_DIR", "/tmp/work")
FINANCIAL_JSON_PATH = os.path.join(BASE_DIR, "financial.json")
EXCEL_PATH = os.path.join(BASE_DIR, "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx")

TARGET_SHEET_KEYWORD = "総合所見"

CELL_MAP = {
    "A7": "profitability",
    "A17": "funding",
    "A27": "stability",
    "A37": "other",
    "A47": "summary",   # 追加修正1
}

AUTO_HEIGHT_CELLS = ["A7", "A17", "A27", "A37", "A47"]

MODEL = "gpt-4.1-mini"


# =========================
# OpenAI API キー設定（Colab依存を除去）
# =========================

api_key = os.environ.get("OPENAI_API_KEY2") or os.environ.get("OPENAI_API_KEY")
if not api_key:
    raise RuntimeError("環境変数 OPENAI_API_KEY2 または OPENAI_API_KEY が未設定です。")

client = OpenAI(api_key=api_key)


# =========================
# system プロンプト
# =========================

SYSTEM_PROMPT = """あなたは財務分析を専門とするアナリストです。
税理士・金融機関・経営者向けの財務分析レポートを作成します。

数値に基づき、客観的かつ慎重な表現を用いてください。
過度な断定や断言は避け、
「〜と考えられる」「〜傾向が見られる」「〜といえる」
といった表現を基本とします。

与えられたデータ以外の推測や創作は行わないでください。
"""


# =========================
# utility：テキスト正規化
# =========================

def normalize_text(text: str) -> str:
    if not text:
        return ""
    return " ".join(text.replace("\n", " ").split())


# =========================
# financial.json 読み込み（形式を柔軟化）
# =========================

def load_financial_json(path: str) -> List[Dict[str, Any]]:
    """
    期待形式:
      1) {"response":[...]}  (従来)
      2) [...]              (配列のみ)
      3) null               (データ無し)
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if data is None:
        return []
    if isinstance(data, dict) and isinstance(data.get("response"), list):
        return data["response"]
    if isinstance(data, list):
        return data

    raise ValueError("financial.json の形式が不正です。{'response':[...]} / [...] / null を期待します。")


# =========================
# 数値処理
# =========================

def to_number(value):
    if value in ("", None, "N/A"):
        return None
    s = str(value).replace(",", "").replace("%", "")
    try:
        if "." in s:
            return float(s)
        return int(s)
    except Exception:
        return None


def calc_trend(values: List[float]) -> str:
    v = [x for x in values if x is not None]
    if len(v) < 2:
        return "stable"
    if v[-1] > v[0]:
        return "improving"
    if v[-1] < v[0]:
        return "deteriorating"
    return "stable"


# =========================
# 中間 JSON 構築
# =========================

def build_section(financial_data, indicators, section, title):
    metrics = []

    for row in financial_data:
        if row["indicator"] not in indicators:
            continue
        if not row.get("usage", True):
            continue

        values = [
            to_number(row.get("previous_previous_term")),
            to_number(row.get("previous_term")),
            to_number(row.get("current_term")),
        ]

        metrics.append({
            "indicator": row["indicator"],
            "values": values,
            "trend": calc_trend(values)
        })

    return {
        "section": section,
        "title": title,
        "periods": ["前々期", "前期", "今期"],
        "metrics": metrics
    }


def build_all_sections(financial_data):
    return {
        "profitability": build_section(
            financial_data,
            ["売上高", "売上総利益率", "営業利益率", "経常利益率", "総資本回転率", "総資産利益率（ROA）"],
            "profitability",
            "収益面"
        ),
        "funding": build_section(
            financial_data,
            ["現預金", "棚卸資産", "棚卸回転率（回転/年）", "流動比率", "当座比率", "借入金", "支払利息"],
            "funding",
            "資金面"
        ),
        "stability": build_section(
            financial_data,
            ["自己資本比率", "負債比率", "財務レバレッジ比率", "固定比率", "固定長期適合率"],
            "stability",
            "財務バランス・安全性"
        ),
        "other": build_section(
            financial_data,
            ["売上高", "営業利益", "総資本回転率", "総資産利益率（ROA）", "有形固定資産額"],
            "other",
            "その他"
        ),
    }


# =========================
# ChatGPT 呼び出し
# =========================

def call_chatgpt(system_prompt, user_prompt):
    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=0.2,
    )
    return normalize_text(response.choices[0].message.content.strip())


# =========================
# 行高さ自動調整
# =========================

def adjust_row_height(ws: Worksheet, cell_address: str):
    cell = ws[cell_address]
    text = cell.value or ""

    cell.alignment = Alignment(wrap_text=True, vertical="top")

    base_height = 15
    chars_per_line = 40

    lines = max(1, math.ceil(len(text) / chars_per_line))
    ws.row_dimensions[cell.row].height = max(base_height, base_height * lines)


# =========================
# Excel 書き込み
# =========================

def write_to_excel(path: str, results: Dict[str, str]):
    wb = openpyxl.load_workbook(path)

    target_sheet_name = None
    for name in wb.sheetnames:
        if TARGET_SHEET_KEYWORD in name:
            target_sheet_name = name
            break

    if not target_sheet_name:
        raise ValueError(f"'{TARGET_SHEET_KEYWORD}' を含むシートが見つかりません。")

    ws: Worksheet = wb[target_sheet_name]

    for cell, key in CELL_MAP.items():
        ws[cell].value = results.get(key, "")
        adjust_row_height(ws, cell)

    wb.save(path)


# =========================
# メイン
# =========================

def main():
    financial_data = load_financial_json(FINANCIAL_JSON_PATH)

    # データ無しなら空欄を書いて終了（最低限の互換）
    if not financial_data:
        write_to_excel(EXCEL_PATH, {
            "profitability": "",
            "funding": "",
            "stability": "",
            "other": "",
            "summary": ""
        })
        print("DONE: financial_data が空のため、空欄で出力しました。")
        return

    sections = build_all_sections(financial_data)
    results = {}

    for key in ["profitability", "funding", "stability", "other"]:
        prompt = f"""以下は企業の「{sections[key]['title']}」に関する財務データ（3期分）です。

{json.dumps(sections[key], ensure_ascii=False, indent=2)}

分析コメントを作成してください。
"""
        results[key] = call_chatgpt(SYSTEM_PROMPT, prompt)

    summary_prompt = f"""以下は財務分析結果の要約です。

【収益面】
{results["profitability"]}

【資金面】
{results["funding"]}

【財務バランス・安全性】
{results["stability"]}

【その他】
{results["other"]}

これらを踏まえ、総合所見を作成してください。
"""

    results["summary"] = call_chatgpt(SYSTEM_PROMPT, summary_prompt)

    write_to_excel(EXCEL_PATH, results)

    print("DONE: 書込み位置・行高さ調整を含めて正常に完了しました。")


if __name__ == "__main__":
    main()
