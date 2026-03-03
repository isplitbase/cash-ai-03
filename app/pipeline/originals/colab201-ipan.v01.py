# -*- coding: utf-8 -*-
"""
colab201-ipan.py
最小版 colab201.py をベースに、以下機能を追加:
  - LibreOffice（headless）で再計算
  - 全シートで「数式 → 値貼り（結果値で置換）」

前提（WORK_DIR 配下）:
  - エクセル転記仕様.xlsx
  - output_updated.json
  - CF付財務分析表（経営指標あり）_ReadingData.xlsx

出力（WORK_DIR 配下）:
  - CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx
  - transfer_log.txt
"""

import json
import re
import math
import os
import sys
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


# ==========
# パス設定（cash-ai-03 runner の前提に合わせる）
# ==========
BASE_DIR = Path(os.environ.get("WORK_DIR", "/tmp/work")).resolve()

SPEC_PATH = BASE_DIR / "エクセル転記仕様.xlsx"
SRC_EXCEL_PATH = BASE_DIR / "CF付財務分析表（経営指標あり）_ReadingData.xlsx"
JSON_PATH = BASE_DIR / "output_updated.json"

OUT_EXCEL_PATH = BASE_DIR / "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx"
OUT_LOG_PATH = BASE_DIR / "transfer_log.txt"

TARGET_SHEET_NAME = "財務諸表（入力）"


# ==========
# 補助関数
# ==========

def parse_row_set(expr: str) -> set[int]:
    """「転記行」のパース（例: '6-10,12'）"""
    if expr is None:
        return set()
    s = str(expr).strip().replace(" ", "").replace("　", "")
    if not s:
        return set()
    parts = [p for p in s.split(",") if p]
    rows: set[int] = set()
    for p in parts:
        if "-" in p:
            a, b = p.split("-", 1)
            rows.update(range(int(a), int(b) + 1))
        else:
            rows.add(int(p))
    return rows


def coerce_value(v):
    """JSONの値をExcel用に変換（数値変換・カンマ除去）"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s2 = s.replace(",", "").replace("，", "")
        # (123) -> -123 対応
        m = re.fullmatch(r"\(([-+]?\d+(\.\d+)?)\)", s2)
        if m:
            s2 = "-" + m.group(1)
        try:
            if re.fullmatch(r"[-+]?\d+", s2):
                return int(s2)
            return float(s2)
        except Exception:
            return s
    return v


def writable_cell(ws, row: int, col_letter: str):
    """結合セルの場合、左上（アンカー）セルを返す"""
    col_idx = column_index_from_string(col_letter)
    coord = f"{col_letter}{row}"
    for mr in ws.merged_cells.ranges:
        if coord in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col)
    return ws.cell(row=row, column=col_idx)


def is_merged_child(ws, coord: str) -> bool:
    """coord が結合セルの“左上以外”なら True"""
    for mr in ws.merged_cells.ranges:
        if coord in mr:
            top_left = f"{get_column_letter(mr.min_col)}{mr.min_row}"
            return coord != top_left
    return False


def run_libreoffice_recalc_xlsx(input_xlsx: Path, out_dir: Path, log_lines: list[str]) -> Path:
    """
    LibreOffice headless でファイルを開いて変換（保存）することで再計算を反映させる。
    生成物は out_dir に出る（basename同名 .xlsx）。

    失敗時は例外。
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError("LibreOffice が見つかりません（soffice/libreoffice コマンドが必要です）")

    out_dir.mkdir(parents=True, exist_ok=True)

    # LibreOffice変換: xlsx -> xlsx（Calc）として出力
    # filter名は環境差があるため、まずは一般的な指定で実行
    cmd = [
        soffice,
        "--headless",
        "--nologo",
        "--nolockcheck",
        "--nodefault",
        "--norestore",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(out_dir),
        str(input_xlsx),
    ]

    log_lines.append(f"[LO] run: {' '.join(cmd)}")

    p = subprocess.run(cmd, capture_output=True, text=True)
    log_lines.append(f"[LO] returncode={p.returncode}")
    if p.stdout.strip():
        log_lines.append(f"[LO] stdout: {p.stdout.strip()}")
    if p.stderr.strip():
        log_lines.append(f"[LO] stderr: {p.stderr.strip()}")

    if p.returncode != 0:
        raise RuntimeError("LibreOffice 変換（再計算）に失敗しました。stderr を確認してください。")

    out_path = out_dir / input_xlsx.name
    if not out_path.exists():
        # まれに拡張子やbasenameが変わるケースに備えて探索
        candidates = list(out_dir.glob(input_xlsx.stem + "*.xlsx"))
        if len(candidates) == 1:
            out_path = candidates[0]
        else:
            raise RuntimeError(f"LibreOffice 変換結果が見つかりません: {out_path}")

    return out_path


def paste_values_for_all_formulas(xlsx_with_formulas: Path, xlsx_evaluated: Path, out_path: Path, log_lines: list[str]) -> None:
    """
    xlsx_with_formulas: 数式/書式を保持したブック（data_only=False）
    xlsx_evaluated:     LibreOffice再計算済み（data_only=True で値が取れる想定）
    すべてのシート・すべての数式セルを「結果値」に置換して保存する。
    """
    wb_f = openpyxl.load_workbook(str(xlsx_with_formulas), data_only=False)
    wb_v = openpyxl.load_workbook(str(xlsx_evaluated), data_only=True)

    # 念のため計算フラグ
    wb_f.calculation.calcMode = "auto"
    wb_f.calculation.fullCalcOnLoad = True

    replaced = 0
    skipped_merged_children = 0

    for sheet_name in wb_f.sheetnames:
        if sheet_name not in wb_v.sheetnames:
            log_lines.append(f"[PASTE] skip sheet not in evaluated: {sheet_name}")
            continue

        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name]

        # 走査（max_row/max_col は重いので iter_rows を使う）
        for row in ws_f.iter_rows():
            for cell_f in row:
                val_f = cell_f.value
                if not (isinstance(val_f, str) and val_f.startswith("=")):
                    continue

                coord = cell_f.coordinate
                if is_merged_child(ws_f, coord):
                    skipped_merged_children += 1
                    continue

                cell_v = ws_v[coord]
                # 数式の結果が None の場合も「値貼り」目的なので None を入れる（空にする）
                cell_f.value = cell_v.value
                replaced += 1

    wb_f.save(str(out_path))
    log_lines.append(f"[PASTE] replaced_formula_cells={replaced}, skipped_merged_children={skipped_merged_children}")


# ==========
# メイン処理
# ==========

def main() -> int:
    log_lines = [f"--- Transfer Log {datetime.now()} ---"]
    log_lines.append(f"WORK_DIR={BASE_DIR}")

    # 0) ファイル存在チェック
    for p in [SPEC_PATH, SRC_EXCEL_PATH, JSON_PATH]:
        if not p.exists():
            print(f"エラー: ファイルが見つかりません -> {p}", file=sys.stderr)
            log_lines.append(f"[ERROR] missing: {p}")
            OUT_LOG_PATH.write_text("\n".join(log_lines), encoding="utf-8")
            return 1

    # 1) 転記仕様の読み取り
    spec_wb = openpyxl.load_workbook(str(SPEC_PATH), data_only=True)
    sheet_name = "ルール(正)" if "ルール(正)" in spec_wb.sheetnames else spec_wb.sheetnames[0]
    rules_ws = spec_wb[sheet_name]

    rules = []
    for r in range(3, rules_ws.max_row + 1):
        target = rules_ws.cell(r, 1).value
        key_name = rules_ws.cell(r, 2).value
        col = rules_ws.cell(r, 3).value
        row_expr = rules_ws.cell(r, 4).value
        if not all([target, key_name, col, row_expr]):
            continue

        rules.append({
            "key_name": str(key_name),
            "column": str(col).strip(),
            "allowed_rows": parse_row_set(str(row_expr)),
        })

    log_lines.append(f"[SPEC] rules={len(rules)} sheet={sheet_name}")

    # 2) JSON読み取り
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        records = json.load(f)
    log_lines.append(f"[JSON] records={len(records) if isinstance(records, list) else 'not_list'}")

    # 3) Excel読み取り（数式保持）
    wb = openpyxl.load_workbook(str(SRC_EXCEL_PATH), data_only=False)
    if TARGET_SHEET_NAME not in wb.sheetnames:
        print(f"エラー: シート「{TARGET_SHEET_NAME}」がありません。", file=sys.stderr)
        log_lines.append(f"[ERROR] missing sheet: {TARGET_SHEET_NAME}")
        OUT_LOG_PATH.write_text("\n".join(log_lines), encoding="utf-8")
        return 1
    ws = wb[TARGET_SHEET_NAME]

    # 再計算フラグ（Excel/LOが開いたときに再計算しやすくする）
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # 4) 転記実行
    stats = {"cells_written": 0, "rows_written": 0}

    SPEC_TO_JSON_KEY = {
        "category": "区分",
        "account_name": "勘定科目",
        "value_t-2": "前々期",
        "value_t-1": "前期",
        "value_t": "今期",
        "aggregation_method": "集計方法",
    }

    for rec in records if isinstance(records, list) else []:
        cell_row_raw = rec.get("セル")
        if cell_row_raw is None:
            continue
        try:
            excel_row = int(str(cell_row_raw).strip())
        except Exception:
            continue

        account_val = coerce_value(rec.get("勘定科目"))
        account_is_blank = (account_val is None or str(account_val).strip() == "")

        wrote_any = False
        for rule in rules:
            if excel_row not in rule["allowed_rows"]:
                continue

            json_key = SPEC_TO_JSON_KEY.get(rule["key_name"], rule["key_name"])

            if (json_key == "集計方法" or json_key == "備考") and account_is_blank:
                value = None
            else:
                value = coerce_value(rec.get(json_key))

            cell = writable_cell(ws, excel_row, rule["column"])
            cell.value = value
            stats["cells_written"] += 1
            wrote_any = True

        if wrote_any:
            stats["rows_written"] += 1

    log_lines.append(f"[WRITE] {stats}")

    # 5) まず一旦保存（数式は残った状態の updated.xlsx）
    #    このファイルを LibreOffice で開いて再計算・保存させる
    tmp_formula_path = OUT_EXCEL_PATH
    wb.save(str(tmp_formula_path))
    log_lines.append(f"[SAVE] formula_kept: {tmp_formula_path}")

    # 6) LibreOfficeで再計算（headless）
    with tempfile.TemporaryDirectory(prefix="lo_recalc_", dir=str(BASE_DIR)) as td:
        lo_out_dir = Path(td)
        evaluated_path = run_libreoffice_recalc_xlsx(tmp_formula_path, lo_out_dir, log_lines)
        log_lines.append(f"[LO] evaluated: {evaluated_path}")

        # 7) 全シートで「数式→値貼り」
        #    ここで OUT_EXCEL_PATH を “値貼り済み” で上書き保存する
        paste_values_for_all_formulas(
            xlsx_with_formulas=tmp_formula_path,
            xlsx_evaluated=evaluated_path,
            out_path=OUT_EXCEL_PATH,
            log_lines=log_lines,
        )
        log_lines.append(f"[SAVE] values_pasted: {OUT_EXCEL_PATH}")

    # 8) ログ出力
    OUT_LOG_PATH.write_text("\n".join(log_lines), encoding="utf-8")

    print("完了しました。")
    print(f"出力ファイル: {OUT_EXCEL_PATH}")
    print(f"書き込みセル数: {stats['cells_written']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
